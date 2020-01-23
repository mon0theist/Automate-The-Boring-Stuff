#! /usr/bin/python
#
# ATBS Chapter 12 - Practice Projects - Text File to Spreadsheet
#
# Write a program to read in the contents of several text files (you can make the
# text files yourself) and insert those contents into a spreadsheet, with one line
# of text per row. The lines of the first text file will be in the cells of column A,
# the lines of the second text file will be in the cells of column B, and so on.
#
# Use the readlines() File object method to return a list of strings, one string
# per line in the file. For the first file, output the first line to column 1,
# row 1. The second line should be written to column 1, row 2, and so on. The next
# file that is read with readlines() will be written to column 2, the next file
# to column 3, and so on.

import openpyxl
from openpyxl.styles import Font

# load/create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active

# open/read files
# Column A
text_file1 = open('file1.txt', 'r')
lines1 = text_file1.readlines()
for i in range(1, len(lines1) + 1):
    ws['A' + str(i)].value = lines1[i - 1]

# Column B
text_file2 = open('file2.txt', 'r')
lines2 = text_file2.readlines()
for i in range(1, len(lines1) + 1):
    ws['B' + str(i)].value = lines2[i - 1]

# Column C
text_file3 = open('file3.txt', 'r')
lines3 = text_file3.readlines()
for i in range(1, len(lines1) + 1):
    ws['C' + str(i)].value = lines3[i - 1]

# this could probably be done all in one loop instead of 3 separate ones, but
# whatever it still works ;)

wb.save('text_to_spreadsheet.xlsx')
