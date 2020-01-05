#! /usr/bin/python
#
# ATBS Chapter 12 - Updating a Spreadsheet
# produceSales.xlsx
#
# Update the following prices:
# Celery - 1.19
# Garlic - 3.07
# Lemon - 1.27

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
# If any future price updates are required, you can just update this dictionary
# whereas a hard-coded solution would require you to change the whole code, which
# could introduce new bugs
PRICE_UPDATES = {
'Garlic': 3.07,
'Celery': 1.19,
'Lemon': 1.27
}

# TODO: Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row): # skip the first row since first row is column names
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updated_produce_sales.xlsx')
