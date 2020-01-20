#! /usr/bin/python
#
# ATBS Chapter 12 - Practice Projects - Blank Row Inserter
#
# Create a program that takes two integers and a filename string as command line
# arguments. Let’s call the first integer N and the second integer M. Starting at
# row N, the program should insert M blank rows into the spreadsheet. For example,
# when the program is run like this:
#
# python blankRowInserter.py 3 2 myProduce.xlsx
#
# the “before” and “after” spreadsheets should look like Figure 12-12.
#
# You can write this program by reading in the contents of the spreadsheet. Then,
# when writing out the new spreadsheet, use a for loop to copy the first N rows.
# For the remaining rows, add M to the row number in the output spreadsheet.

import sys, os, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

n = int(sys.argv[1]) # row, insert blank rows after this row
m = int(sys.argv[2]) # number of blank rows
# file = str(sys.argv[3])

# load workbook(s), set sheets
wb_old = openpyxl.load_workbook(sys.argv[3])
sheet_old = wb_old.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

# Validate input/args
if n > sheet_old.max_row:
    print('WARNING: Insertion point value (sys.argv[1]) is greater than total number of rows. Program will still work, but will have no practical effect.\n')

print('\nProcessing ' + str(sheet_old.max_row + m) + ' total rows...\n')
print('If processing is taking too long, hit CTRL+C and try again with a smaller spreadsheet')

try:
    # copy rows to new sheet
    for i in range(1, (n + 1)):
        for j in range(1, (sheet_old.max_column + 1)):
            sheet_new[get_column_letter(j) + str(i)].value = sheet_old[get_column_letter(j) + str(i)].value

    # insert blank rows into new sheet
    # blank rows from n+1 to n+1+m
    for i in range(n+1, n+1+m+1):
        for j in range(1, (sheet_old.max_column + 1)):
            sheet_new[get_column_letter(j) + str(i)].value = ""

    # copy rest of rows after blank rows
    for i in range(n+1+m, (sheet_old.max_row + 1)):
        for j in range(1, (sheet_old.max_column + 1)):
            sheet_new[get_column_letter(j) + str(i + m)].value = sheet_old[get_column_letter(j) + str(i)].value
except KeyboardInterrupt:
    print('\nKeyboard Interrupt by user\n')
    exit()

print('Saving to file: blank_rows_inserted.xlsx')
wb_new.save('blank_rows_inserted.xlsx')

print('Done!')
