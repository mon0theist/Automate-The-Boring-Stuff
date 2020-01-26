#! /usr/bin/python
#
# ATBS Chapter 12 - Practice Projects - Spreadsheet to Text
#
# Write a program that performs the tasks of the previous program in reverse order:
# The program should open a spreadsheet and write the cells of column A into one
# text file, the cells of column B into another text file, and so on.

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# using workbook from previous program
wb = openpyxl.load_workbook('text_to_spreadsheet.xlsx')
ws = wb.active

# column A
file1 = open('stt1.txt', 'w')
for i in range(1, (ws.max_row + 1)):
    file1.write(ws['A' + str(i)].value)
print('Wrote output to stt1.txt')
file1.close()

# column B
file2 = open('stt2.txt', 'w')
for i in range(1, (ws.max_row + 1)):
    file2.write(ws['B' + str(i)].value)
print('Wrote output to stt2.txt')
file2.close()

# column C
file3 = open('stt3.txt', 'w')
for i in range(1, (ws.max_row + 1)):
    file3.write(ws['C' + str(i)].value)
print('Wrote output to stt3.txt')
file3.close()

# probably could've been done more elegantly in one loop instead of 3 loops but
# whatever, should still work

print('Done!')
