#! /usr/bin/python
#
# ATBS Chapter 12 - Practice Projects - Spreadsheet Cell Inverter
#
# Write a program to invert the row and column of the cells in the spreadsheet.
# For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
# This should be done for all cells in the spreadsheet. For example, the “before”
# and “after” spreadsheets would look something like Figure 12-13.
#
# You can write this program by using nested For Loops to read the spreadsheet’s
# data into a list of lists data structure. This data structure could have
# sheetData[x][y] for the cell at column x and row y. Then, when writing out the
# new spreadsheet, use sheetData[y][x] for the cell at column x and row y.

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# load workbooks/worksheets
wb = openpyxl.load_workbook('invert_me.xlsx')
print('Loaded workbook: invert_me.xlsx')
ws = wb.active
wb2 = openpyxl.Workbook()
ws2 = wb2.active

# gather spreadsheet data
sheet_data = []
for i in range(1, ws.max_column + 1):
    col_list = []
    for j in range(1, ws.max_row + 1):
        cell = get_column_letter(i) + str(j)
        col_list.append(ws[cell].value)
    sheet_data.append(col_list)

# write new inverted spreadsheet
for x in range(1, ws.max_column + 1):
    for y in range(1, ws.max_row + 1):
        ws2[get_column_letter(y) + str(x)].value = sheet_data[(x - 1)][(y - 1)]

print('Saving file: invert_me_output.xlsx')
wb2.save('invert_me_output.xlsx')

print('Done!')
