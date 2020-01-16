#! /usr/bin/python
#
# ATBS Chapter 12 - Practice Projects - Multiplication Table Maker
#
# Create a program multiplicationTable.py that takes a number N from the command
# line and creates an N×N multiplication table in an Excel spreadsheet.
# For example, when the program is run like this:
#
# py multiplicationTable.py 6
#
# ... it should create a spreadsheet that looks like Figure 12-11.

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
# DEPRECATED:
# from openpyxl.cell import get_column_letter, column_index_from_string

# Determine n, validate user input
try:
    n = int(sys.argv[1])

    # Create workbook, set sheet
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Cell A1 should be blank, but this line is probably unnecessary
    sheet.cell(row=1, column=1).value = ''

    # bold font object
    bold_font = Font(bold=True)

    print("Generating multiplication table...")

    # create x-axis (row 1, columns 2 thru n)
    for i in range(2, n+2):
        sheet.cell(row=1, column=i).value = str(i - 1)
        # formatting
        current_cell = sheet.cell(row=1, column=i)
        current_cell.alignment = Alignment(horizontal='center')
        current_cell.font = bold_font

    # create y-axis (column 1/A, rows 2 thru n)
    for i in range(2, n+2):
        sheet.cell(row=i, column=1).value = str(i - 1)
        # formatting
        current_cell = sheet.cell(row=i, column=1)
        current_cell.alignment = Alignment(horizontal='center')
        current_cell.font = bold_font

    # write cells
    for i in range(2, (sheet.max_row + 1)):
        for j in range(2, sheet.max_column + 1):
            formula = "=" + get_column_letter(j) + "1*A" + str(i)
            sheet.cell(row=i, column=j).value = formula
            # formatting
            current_cell = sheet.cell(row=i, column=j)
            current_cell.alignment = Alignment(horizontal='center')

    print('Multiplication table is complete!')
    wb.save('multiplication_table_results.xlsx')
    print('Saved/overwritten to multiplication_table_results.xlsx')

except ValueError:
        print('Invalid argument: Please enter an int value\n')
        exit()
except IndexError:
        print('Invalid argument: You did not enter an argument\n')
        exit()
